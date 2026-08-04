"""Tests for tax-workbook and standalone Excel export modes."""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from excel_exporter import ExcelExporter  # noqa: E402


def _sample_data() -> dict:
    return {
        "accounts": [{"account": "UACCT123"}],
        "account_summary": [],
        "account_values": [],
        "positions": [
            {
                "account": "UACCT123",
                "symbol": "AMZN",
                "instrument_kind": "Stock",
                "security_type": "STK",
                "currency": "USD",
                "exchange": "NASDAQ",
                "primary_exchange": "NASDAQ",
                "contract_id": 1,
                "quantity": 7,
                "average_cost": 1.0,
            },
            {
                "account": "UACCT123",
                "symbol": "EUR",
                "instrument_kind": "FX Cash",
                "security_type": "CASH",
                "currency": "USD",
                "exchange": "IDEALPRO",
                "primary_exchange": "",
                "contract_id": 2,
                "quantity": 1000,
                "average_cost": 0,
            },
        ],
        "portfolio": [],
        "errors": [],
    }


def test_standalone_uses_ibkr_prefix(tmp_path: Path) -> None:
    out = tmp_path / "IBKR_Portfolio.xlsx"
    ExcelExporter(out, mode="standalone").export(_sample_data())
    wb = load_workbook(out)
    assert "IBKR_Positions" in wb.sheetnames
    assert "Positions" not in wb.sheetnames
    assert "IBKR_Reconciliacao" not in wb.sheetnames
    assert "IBKR_Overview" in wb.sheetnames


def test_tax_mode_preserves_foreign_and_writes_recon(tmp_path: Path) -> None:
    path = tmp_path / "tax.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MyProfit_2026"
    ws["A4"] = "Nome"
    ws["B4"] = "Ativo normalizado"
    ws["C4"] = "Quantidade"
    ws["B5"] = "AMZN"
    ws["C5"] = 7
    ws["B6"] = "BIL"
    ws["C6"] = 100
    wb.create_sheet("Registro_Real")["A1"] = "keep-me"
    wb.save(path)

    ExcelExporter(path, mode="tax_workbook", qty_tolerance=0.0001).export(_sample_data())
    wb2 = load_workbook(path)
    assert wb2["Registro_Real"]["A1"].value == "keep-me"
    assert "IBKR_Positions" in wb2.sheetnames
    assert "IBKR_Reconciliacao" in wb2.sheetnames

    statuses: dict[str, str] = {}
    sheet = wb2["IBKR_Reconciliacao"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            statuses[str(row[0])] = str(row[4])
    assert statuses["AMZN"] == "OK"
    assert statuses["BIL"] == "ONLY_FISCAL"
    assert "EUR" not in statuses


def test_standalone_overview_has_refresh_control(tmp_path: Path) -> None:
    out = tmp_path / "IBKR_Portfolio.xlsx"
    ExcelExporter(out, mode="standalone").export(_sample_data())
    overview = load_workbook(out)["IBKR_Overview"]
    values = {
        overview.cell(row=r, column=1).value: overview.cell(row=r, column=2).value
        for r in range(1, overview.max_row + 1)
        if overview.cell(row=r, column=1).value
    }
    assert "Update data" in values
    # No repo scripts/ next to tmp_path → fallback label, not a hyperlink target.
    assert values["Update data"] == "Run ./scripts/refresh_workbook.command"


def test_overview_links_refresh_command_when_present(tmp_path: Path) -> None:
    project = tmp_path / "proj"
    scripts = project / "scripts"
    scripts.mkdir(parents=True)
    command = scripts / "refresh_workbook.command"
    command.write_text("#!/bin/zsh\n", encoding="utf-8")
    out = project / "data" / "output" / "IBKR_Portfolio.xlsx"
    out.parent.mkdir(parents=True)

    ExcelExporter(out, mode="standalone").export(_sample_data())
    overview = load_workbook(out)["IBKR_Overview"]
    action = None
    for row in overview.iter_rows(min_row=1, max_col=2):
        if row[0].value == "Update data":
            action = row[1]
            break
    assert action is not None
    assert action.value == "Refresh from TWS"
    assert action.hyperlink is not None
    assert action.hyperlink.target == command.resolve().as_uri()


def test_tax_mode_missing_file_raises(tmp_path: Path) -> None:
    missing = tmp_path / "nope.xlsx"
    with pytest.raises(FileNotFoundError):
        ExcelExporter(missing, mode="tax_workbook").export(_sample_data())


def test_tax_mode_removes_legacy_owned_sheets(tmp_path: Path) -> None:
    path = tmp_path / "tax.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MyProfit_2026"
    ws["B4"] = "Ativo normalizado"
    ws["C4"] = "Quantidade"
    wb.create_sheet("Positions")["A1"] = "legacy"
    wb.create_sheet("Overview")["A1"] = "legacy"
    wb.save(path)

    ExcelExporter(path, mode="tax_workbook").export(_sample_data())
    names = load_workbook(path).sheetnames
    assert "Positions" not in names
    assert "Overview" not in names
    assert "IBKR_Positions" in names
