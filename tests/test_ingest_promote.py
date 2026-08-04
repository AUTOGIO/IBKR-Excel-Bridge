"""Tests for statement ingest and Registro_Real promote."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from events_store import load_events, merge_events, save_events  # noqa: E402
from ingest_statements import load_aliases, parse_statement_csv  # noqa: E402
from promote_events import (  # noqa: E402
    derive_positions_from_events,
    promote_events_to_workbook,
    write_staging_sheets,
)

FIXTURES = Path(__file__).resolve().parent / "fixtures"


def test_parse_trades_buy_sell() -> None:
    events = parse_statement_csv(FIXTURES / "sample_trades.csv")
    assert len(events) == 2
    by_sym = {e["symbol"]: e for e in events}
    assert by_sym["BIL"]["tipo_evento"] == "Compra"
    assert by_sym["BIL"]["quantity"] == 100
    assert by_sym["BIL"]["date"] == "2026-08-01"
    assert by_sym["AAPL"]["tipo_evento"] == "Venda"
    assert by_sym["AAPL"]["quantity"] == 5


def test_parse_dividend_and_alias() -> None:
    aliases = {"STAGX": "STAG"}
    # rewrite? use STAG in fixture; test alias with trades
    events = parse_statement_csv(FIXTURES / "sample_dividends.csv", aliases=aliases)
    assert len(events) == 1
    assert events[0]["tipo_evento"] == "Rendimento"
    assert events[0]["ir_retido_usd"] == 11.51


def test_merge_dedupes_and_preserves_promoted(tmp_path: Path) -> None:
    events = parse_statement_csv(FIXTURES / "sample_trades.csv")
    path = tmp_path / "events.jsonl"
    save_events(path, events)
    again, new_count = merge_events(load_events(path), events)
    assert new_count == 0
    assert len(again) == 2
    again[0]["promoted"] = True
    merged, new_count = merge_events(again, events)
    assert new_count == 0
    assert any(e.get("promoted") for e in merged)


def test_promote_appends_formulas_idempotent(tmp_path: Path) -> None:
    wb_path = tmp_path / "tax.xlsx"
    events_path = tmp_path / "events.jsonl"
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro_Real"
    headers = [
        "Data",
        "Ativo",
        "Tipo_Evento",
        "Quantidade",
        "Preco_Unit_USD",
        "PTAX",
        "IR_Retido_USD",
        "Valor_Total_BRL",
        "Custo_Medio_BRL_Vigente",
        "Resultado_BRL",
        "IR_Retido_BRL",
        "Observacoes",
        "Status_Validacao",
        "Credito_Exterior_Limitado_BRL",
        "Retencao_Exterior_Nao_Aproveitada_BRL",
    ]
    for idx, h in enumerate(headers, start=1):
        ws.cell(4, idx, value=h)
    # seed template row 5
    ws["A5"] = "2026-01-01"
    ws["B5"] = "SEED"
    ws["C5"] = "Compra"
    ws["D5"] = 1
    ws["E5"] = 10
    ws["F5"] = 5
    ws["H5"] = '=IF(D5="","",IF(OR(E5="",F5=""),"",D5*E5*F5))'
    ws["I5"] = '=IF(B5="","",1)'
    ws["J5"] = '=IF(C5="Venda",1,"")'
    ws["K5"] = '=IF(C5="Rendimento",1,0)'
    ws["M5"] = '=IF(A5="","","OK")'
    ws["N5"] = "=0"
    ws["O5"] = "=0"
    wb.create_sheet("MyProfit_2026")
    wb.save(wb_path)

    events = parse_statement_csv(FIXTURES / "sample_trades.csv")
    save_events(events_path, events)

    summary = promote_events_to_workbook(wb_path, events_path)
    assert summary["appended"] == 2
    ws2 = load_workbook(wb_path)["Registro_Real"]
    assert ws2["B6"].value == "AAPL" or ws2["B7"].value == "AAPL"
    # formulas retargeted
    assert (
        "D6" in str(ws2["H6"].value) or "D7" in str(ws2["H6"].value) or "D6" in str(ws2["H7"].value)
    )

    summary2 = promote_events_to_workbook(wb_path, events_path)
    assert summary2["appended"] == 0
    assert all(e.get("promoted") for e in load_events(events_path))


def test_derive_positions() -> None:
    events = parse_statement_csv(FIXTURES / "sample_trades.csv")
    rows = {r["symbol"]: r for r in derive_positions_from_events(events)}
    assert "BIL" in rows
    assert rows["BIL"]["quantity"] == 100


def test_write_staging_sheets(tmp_path: Path) -> None:
    wb_path = tmp_path / "tax.xlsx"
    wb = Workbook()
    wb.active.title = "Registro_Real"
    wb.save(wb_path)
    events = parse_statement_csv(FIXTURES / "sample_trades.csv")
    write_staging_sheets(wb_path, events)
    names = load_workbook(wb_path).sheetnames
    assert "IBKR_Eventos_Staging" in names
    assert "IBKR_Posicao_From_Events" in names


def test_load_aliases(tmp_path: Path) -> None:
    path = tmp_path / "aliases.json"
    path.write_text('{"LQDEz": "LQDE"}', encoding="utf-8")
    aliases = load_aliases(path)
    assert aliases["LQDEZ"] == "LQDE"


def test_map_tipo_requires_explicit_signal(tmp_path: Path) -> None:
    """Rows with only a signed quantity (corporate actions, transfers, FX
    conversions) must be dropped, not silently classified as trades."""
    csv = tmp_path / "ambiguous.csv"
    csv.write_text(
        "Symbol,Date/Time,Quantity,TradePrice,Type\n"
        # No Buy/Sell column and Type is neither BUY/SELL/DIVIDEND.
        "FOO,20260710;000000,100,10.00,CorporateAction\n"
        # Explicit BUY still parses through.
        "BAR,20260710;000000,50,20.00,BUY\n",
        encoding="utf-8",
    )
    events = parse_statement_csv(csv)
    symbols = {e["symbol"] for e in events}
    assert symbols == {"BAR"}, f"FOO should have been dropped: {events}"
