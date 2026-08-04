"""Excel workbook exporter for the IBKR read-only snapshot.

Produces machine-owned ``IBKR_*`` tabs. In ``tax_workbook`` mode, preserves
all foreign Lei 14.754 sheets and adds ``IBKR_Reconciliacao``.
"""

from __future__ import annotations

import logging
import zipfile
from collections.abc import Iterable
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from reconcile import reconcile_quantities

LOGGER = logging.getLogger(__name__)

# openpyxl raises these when a file is truly unreadable as XLSX. Any other
# exception (PermissionError, OSError, …) should surface to the caller
# instead of being silently replaced with a fresh workbook.
_CORRUPT_XLSX_EXCEPTIONS = (zipfile.BadZipFile, InvalidFileException)


# Columns rendered with the "money" format (2 decimals, thousands separator).
_MONEY_COLUMNS: frozenset[str] = frozenset(
    {
        "average_cost",
        "market_price",
        "market_value",
        "unrealized_pnl",
        "realized_pnl",
        "value_numeric",
    }
)

# Columns rendered as quantities (up to 4 decimals to accommodate fractional
# shares and FX lots).
_QUANTITY_COLUMNS: frozenset[str] = frozenset({"quantity", "qty_ibkr", "qty_fiscal", "delta"})

_MONEY_FORMAT: str = "#,##0.00;[Red]-#,##0.00"
_QUANTITY_FORMAT: str = "#,##0.####"

# Default Calibri is 11pt; raise all exporter text by +2 for readability.
_BASE_FONT_SIZE: int = 13
_TITLE_FONT_SIZE: int = 18
_BODY_FONT: Font = Font(size=_BASE_FONT_SIZE)
_HEADER_FONT: Font = Font(size=_BASE_FONT_SIZE, bold=True, color="FFFFFFFF")
_EMPTY_FONT: Font = Font(size=_BASE_FONT_SIZE, bold=True, italic=True)
_TITLE_FONT: Font = Font(size=_TITLE_FONT_SIZE, bold=True)
_GUIDE_LABEL_FONT: Font = Font(size=_BASE_FONT_SIZE, bold=True, color="FF1F4E78")
_GUIDE_BODY_FONT: Font = Font(size=_BASE_FONT_SIZE, italic=True)
_GUIDE_FILL: PatternFill = PatternFill(
    start_color="FFE8F1F8", end_color="FFE8F1F8", fill_type="solid"
)
_REFRESH_FILL: PatternFill = PatternFill(
    start_color="FF1F7A4D", end_color="FF1F7A4D", fill_type="solid"
)
_REFRESH_FONT: Font = Font(size=_BASE_FONT_SIZE, bold=True, color="FFFFFFFF")
_REFRESH_LABEL_FONT: Font = Font(size=_BASE_FONT_SIZE, bold=True)

# Guide rows written above each data table: Description / How to read / Tip.
_GUIDE_ROW_COUNT: int = 4  # 3 text rows + 1 blank spacer before the table header

# Shared Positions ↔ Portfolio comparison shown on both tabs.
_POSITIONS_VS_PORTFOLIO: str = (
    "Positions vs Portfolio: Positions (reqPositions) is the full "
    "holdings list — stocks, futures, options, and FX cash — with qty "
    "and price-only average cost. Portfolio (updatePortfolio) is marked "
    "securities only (no FX cash) with market price, market value, and "
    "P&L; average cost is often commission-inclusive. Use Positions for "
    "what you hold; Portfolio for valuation. Expect more Positions rows "
    "when FX Cash exists. Match stocks/futures by Contract ID, not row "
    "order. Settled FX cash truth is IBKR_Cash_By_Currency. Small MV/P&L "
    "drift vs TWS is normal."
)

SHEET_GUIDES: dict[str, dict[str, str]] = {
    "IBKR_Overview": {
        "description": (
            "Snapshot control panel for this refresh: when it ran, which "
            "account(s), exporter mode, and row counts per tab."
        ),
        "how_to_read": (
            "Start here after every refresh. Confirm Generated timestamp is "
            "recent, Accounts matches your TWS paper/live id, and row counts "
            "look sane (Positions ≥ Portfolio when you hold FX)."
        ),
        "tip": (
            "Click Refresh from TWS (green control below), or run "
            "./scripts/refresh_workbook.command. That closes this workbook, "
            "snapshots TWS, and reopens the file. This tab has no market "
            "prices — only metadata."
        ),
    },
    "IBKR_Account_Summary": {
        "description": (
            "Base-currency account health from IBKR reqAccountSummary: "
            "NetLiquidation, cash, buying power, margin, cushion."
        ),
        "how_to_read": (
            "One row per Tag. Use Value (Numeric) for sorting/sums/charts; "
            "Value is the raw IBKR string (needed for non-numeric tags like "
            "AccountType). Currency is usually USD or blank."
        ),
        "tip": (
            "Reconcile NetLiquidation and TotalCashValue against the TWS "
            "Account Window. Cushion ≈ ExcessLiquidity / NetLiquidation. "
            "Do not sum the Value column as text."
        ),
    },
    "IBKR_Cash_By_Currency": {
        "description": (
            "Per-currency ledger from IBKR $LEDGER-* account values: cash, "
            "FX rate, net liq by currency, and currency-scoped P&L."
        ),
        "how_to_read": (
            "Filter Metric = CashBalance to see balances by Currency. "
            "ExchangeRate converts that currency into your base. "
            "NetLiquidationByCurrency × ExchangeRate should roughly rebuild "
            "base NetLiquidation when summed across currencies."
        ),
        "tip": (
            "This is the source of truth for EUR/GBP/JPY cash — not the "
            "Positions FX Cash rows. Settled cash can lag Positions qty "
            "during T+2 FX settlement."
        ),
    },
    "IBKR_Positions": {
        "description": (
            "Contract-level holdings from reqPositions, including stocks, "
            "futures, options, and FX cash-conversion contracts. Sister tab: "
            "IBKR_Portfolio (valuation / P&L, no FX cash)."
        ),
        "how_to_read": (
            "Use Kind to separate Stock / Future / FX Cash. Quantity is "
            "signed (negative = short). Average Cost here is the native "
            "(price-only) average from IBKR positions — not the "
            "commission-inclusive cost on IBKR_Portfolio."
        ),
        "tip": _POSITIONS_VS_PORTFOLIO,
    },
    "IBKR_Portfolio": {
        "description": (
            "Marked securities from updatePortfolio: market price, market "
            "value, and unrealized/realized P&L. Excludes FX cash contracts. "
            "Sister tab: IBKR_Positions (full holdings list, includes FX)."
        ),
        "how_to_read": (
            "Quantity × Market Price ≈ Market Value. "
            "(Market Price − Average Cost) × Quantity ≈ Unrealized P&L. "
            "Average Cost here is commission-inclusive (can differ slightly "
            "from Positions)."
        ),
        "tip": _POSITIONS_VS_PORTFOLIO,
    },
    "IBKR_API_Messages": {
        "description": (
            "IBKR API status stream for this session: farm connections, "
            "warnings, and real client errors."
        ),
        "how_to_read": (
            "Kind = info is normal (market-data farm OK). Kind = warning "
            "needs a glance. Kind = error (codes 100–499) means a request "
            "failed. Prefer Error Time (Local) over Error Time (ms)."
        ),
        "tip": (
            "A healthy refresh is all info and zero errors. Repeated code "
            "200/502 usually means TWS API socket/port or contract issues."
        ),
    },
    "IBKR_Reconciliacao": {
        "description": (
            "Compares live IBKR position quantities to fiscal quantities "
            "from the tax workbook (MyProfit_2026 / Posicoes_Atuais)."
        ),
        "how_to_read": (
            "Qty IBKR vs Qty Fiscal; Delta = IBKR − Fiscal. Status flags "
            "matches and breaks within the configured tolerance."
        ),
        "tip": (
            "Only available in tax_workbook mode. Investigate BREAK rows "
            "before filing — they often mean missing corporate actions or "
            "unpromoted events."
        ),
    },
    "IBKR_Eventos_Staging": {
        "description": (
            "Staging table of ledger events (trades, dividends, etc.) "
            "pending or already promoted into the tax flow."
        ),
        "how_to_read": (
            "Sort by date/symbol. Promoted = True means the event already "
            "fed derived positions. Check source_file and observacoes for "
            "provenance."
        ),
        "tip": (
            "Do not edit IBKR-owned columns by hand in a way that fights "
            "the next ingest — fix upstream events.jsonl / statements."
        ),
    },
    "IBKR_Posicao_From_Events": {
        "description": (
            "Positions derived by replaying promoted events (quantity and average cost in USD/BRL)."
        ),
        "how_to_read": (
            "Compare Symbol/Quantity here to IBKR_Positions and to fiscal "
            "tabs. Status explains whether the derived lot is open/closed."
        ),
        "tip": (
            "If this diverges from live IBKR Positions, events are incomplete "
            "or promotion rules need review — use IBKR_Reconciliacao next."
        ),
    },
    "Workbook Report": {
        "description": (
            "User-authored audit / structure report. Preserved across "
            "collector refreshes (not overwritten by IBKR_* exporters)."
        ),
        "how_to_read": (
            "Treat findings as guidance for the IBKR_* tabs. Re-check claims "
            "against the latest Generated timestamp on IBKR_Overview."
        ),
        "tip": (
            "After big schema changes, refresh this report manually so it "
            "does not describe a stale layout."
        ),
    },
}

OWNED_SHEETS_STANDALONE: tuple[str, ...] = (
    "IBKR_Overview",
    "IBKR_Account_Summary",
    "IBKR_Cash_By_Currency",
    "IBKR_Positions",
    "IBKR_Portfolio",
    "IBKR_API_Messages",
)

OWNED_SHEETS_TAX: tuple[str, ...] = OWNED_SHEETS_STANDALONE + (
    "IBKR_Reconciliacao",
    "IBKR_Eventos_Staging",
    "IBKR_Posicao_From_Events",
)

LEGACY_OWNED_SHEETS: tuple[str, ...] = (
    "Overview",
    "Account Summary",
    "Cash By Currency",
    "Positions",
    "Portfolio",
    "API Messages",
)


def extract_fiscal_quantities(workbook_path: Path) -> list[dict[str, Any]]:
    """Read fiscal quantities from MyProfit_2026 or Posicoes_Atuais."""
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if "MyProfit_2026" in wb.sheetnames:
            ws = wb["MyProfit_2026"]
            rows: list[dict[str, Any]] = []
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row or len(row) < 3:
                    continue
                symbol, qty = row[1], row[2]
                if symbol is None or str(symbol).strip() == "":
                    continue
                rows.append({"symbol": symbol, "quantity": qty})
            return rows

        if "Posicoes_Atuais" in wb.sheetnames:
            ws = wb["Posicoes_Atuais"]
            rows = []
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row or len(row) < 4:
                    continue
                symbol, qty = row[0], row[3]
                if symbol is None or str(symbol).strip() == "":
                    continue
                rows.append({"symbol": symbol, "quantity": qty})
            return rows

        return []
    finally:
        wb.close()


class ExcelExporter:
    def __init__(
        self,
        output_path: Path,
        *,
        mode: str = "standalone",
        qty_tolerance: float = 0.0001,
        events_file: Path | None = None,
    ) -> None:
        if mode not in {"standalone", "tax_workbook"}:
            raise ValueError(f"Unknown exporter mode: {mode!r}")
        self.output_path = Path(output_path)
        self.mode = mode
        self.qty_tolerance = float(qty_tolerance)
        self.events_file = Path(events_file) if events_file else None
        self.owned_sheets: tuple[str, ...] = (
            OWNED_SHEETS_TAX if mode == "tax_workbook" else OWNED_SHEETS_STANDALONE
        )

    def _assert_not_locked(self) -> None:
        lock = self.output_path.parent / f"~${self.output_path.name}"
        if lock.exists():
            raise PermissionError(
                f"Workbook appears open in Excel ({lock.name}). Close Excel and retry."
            )

    def _load_or_create_workbook(self) -> Workbook:
        """Open an existing workbook (preserving foreign sheets) or create
        a fresh one in standalone mode.

        Sheet order in the returned workbook is:
        - all foreign sheets first (in their original order)
        - then the owned sheets in canonical order, appended by ``export()``.
        """
        overview_name = "IBKR_Overview"

        if self.mode == "tax_workbook":
            if not self.output_path.exists():
                raise FileNotFoundError(
                    f"Tax workbook not found: {self.output_path}. "
                    "Copy the reconciled Lei 14.754 file to this path first."
                )
            self._assert_not_locked()
            try:
                workbook = load_workbook(self.output_path)
            except Exception as exc:  # noqa: BLE001
                raise ValueError(
                    f"Tax workbook is unreadable or corrupt: {self.output_path}"
                ) from exc
        elif not self.output_path.exists():
            workbook = Workbook()
            active = workbook.active
            if active is not None and active.title in {"Sheet", "Sheet1"}:
                workbook.remove(active)
            workbook.create_sheet(overview_name)
            return workbook
        else:
            self._assert_not_locked()
            try:
                workbook = load_workbook(self.output_path)
            except _CORRUPT_XLSX_EXCEPTIONS as exc:
                LOGGER.warning(
                    "Standalone workbook %s is unreadable (%s); rebuilding from scratch.",
                    self.output_path,
                    exc,
                )
                workbook = Workbook()
                default = workbook.active
                if default is not None:
                    workbook.remove(default)
                workbook.create_sheet(overview_name)
                return workbook

        delete_names = set(self.owned_sheets) | set(LEGACY_OWNED_SHEETS)
        for name in list(workbook.sheetnames):
            if name in delete_names:
                del workbook[name]

        workbook.create_sheet(overview_name)
        return workbook

    # -- helpers --

    @staticmethod
    def _humanize(column_name: str) -> str:
        exact = {
            "value_numeric": "Value (Numeric)",
            "error_time_local": "Error Time (Local)",
            "error_time": "Error Time (ms)",
            "instrument_kind": "Kind",
            "security_type": "SecType",
            "primary_exchange": "Primary Exchange",
            "qty_ibkr": "Qty IBKR",
            "qty_fiscal": "Qty Fiscal",
        }
        if column_name in exact:
            return exact[column_name]
        replacements = {
            "Pnl": "P&L",
            "Id": "ID",
            "Api": "API",
        }
        title = column_name.replace("_", " ").title()
        for source, target in replacements.items():
            title = title.replace(source, target)
        return title

    @staticmethod
    def _autosize(worksheet: Worksheet, cap: int = 40) -> None:
        for column_cells in worksheet.columns:
            first = column_cells[0]
            column_letter = getattr(first, "column_letter", None)
            if column_letter is None:
                # Merged cells (e.g. guide body spanning B:F) have no letter.
                col_idx = getattr(first, "column", None)
                if col_idx is None:
                    continue
                column_letter = get_column_letter(col_idx)
            maximum_length = 0
            for cell in column_cells:
                if getattr(cell, "value", None) is None:
                    continue
                value = str(cell.value)
                # Don't let long wrapped guide text inflate column width.
                if len(value) > maximum_length:
                    maximum_length = min(len(value), cap)
            if maximum_length:
                worksheet.column_dimensions[column_letter].width = min(maximum_length + 3, cap)

    @staticmethod
    def _apply_number_format(cell: Any, column: str) -> None:
        if column in _MONEY_COLUMNS:
            cell.number_format = _MONEY_FORMAT
        elif column in _QUANTITY_COLUMNS:
            cell.number_format = _QUANTITY_FORMAT

    @staticmethod
    def _coerce_number(column: str, value: Any) -> Any:
        if column not in _MONEY_COLUMNS and column not in _QUANTITY_COLUMNS:
            return value
        if isinstance(value, (int, float)):
            return value
        if value is None or value == "":
            return None
        try:
            return float(str(value).replace(",", ""))
        except (TypeError, ValueError):
            return value

    @staticmethod
    def _make_table(
        worksheet: Worksheet,
        table_name: str,
        row_count: int,
        column_count: int,
        *,
        start_row: int = 1,
    ) -> None:
        # row_count is the absolute last row of the table (including header).
        if row_count < start_row + 1 or column_count < 1:
            return
        end_column = get_column_letter(column_count)
        table = Table(
            displayName=table_name,
            ref=f"A{start_row}:{end_column}{row_count}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    def _write_sheet_guide(
        self,
        worksheet: Worksheet,
        sheet_name: str,
    ) -> int:
        """Write Description / How to read / Tip block. Returns first data row."""
        guide = SHEET_GUIDES.get(sheet_name)
        if not guide:
            return 1

        lines = (
            ("Description", guide["description"]),
            ("How to read", guide["how_to_read"]),
            ("Tip", guide["tip"]),
        )
        for index, (label, text) in enumerate(lines, start=1):
            label_cell = worksheet.cell(row=index, column=1, value=label)
            body_cell = worksheet.cell(row=index, column=2, value=text)
            label_cell.font = _GUIDE_LABEL_FONT
            body_cell.font = _GUIDE_BODY_FONT
            body_cell.alignment = Alignment(wrap_text=True, vertical="top")
            label_cell.fill = _GUIDE_FILL
            body_cell.fill = _GUIDE_FILL
            worksheet.merge_cells(start_row=index, start_column=2, end_row=index, end_column=6)
            # Tip rows that carry the Positions↔Portfolio comparison need more height.
            worksheet.row_dimensions[index].height = 72 if label == "Tip" else 40

        worksheet.row_dimensions[_GUIDE_ROW_COUNT].height = 8
        worksheet.column_dimensions["A"].width = 14
        worksheet.column_dimensions["B"].width = 28
        return _GUIDE_ROW_COUNT + 1

    def _write_records(
        self,
        workbook: Workbook,
        sheet_name: str,
        records: list[dict[str, Any]],
        table_name: str,
        preferred_order: Iterable[str] | None = None,
    ) -> None:
        worksheet = workbook.create_sheet(sheet_name)
        data_start = self._write_sheet_guide(worksheet, sheet_name)

        if not records:
            empty = worksheet.cell(
                row=data_start,
                column=1,
                value=f"No {sheet_name.lower()} data returned.",
            )
            empty.font = _EMPTY_FONT
            worksheet.column_dimensions["A"].width = 40
            return

        seen_keys: set[str] = set()
        columns: list[str] = []
        if preferred_order:
            for name in preferred_order:
                if name in records[0]:
                    columns.append(name)
                    seen_keys.add(name)
        for name in records[0].keys():
            if name not in seen_keys:
                columns.append(name)

        header_fill = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
        header_row = data_start

        for index, column in enumerate(columns, start=1):
            cell = worksheet.cell(
                row=header_row,
                column=index,
                value=self._humanize(column),
            )
            cell.font = _HEADER_FONT
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for offset, record in enumerate(records):
            row_index = header_row + 1 + offset
            for column_index, column in enumerate(columns, start=1):
                value = self._coerce_number(column, record.get(column))
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                cell.font = _BODY_FONT
                self._apply_number_format(cell, column)

        last_row = header_row + len(records)
        worksheet.freeze_panes = f"A{header_row + 1}"
        self._make_table(
            worksheet=worksheet,
            table_name=table_name,
            row_count=last_row,
            column_count=len(columns),
            start_row=header_row,
        )
        self._autosize(worksheet)
        # Keep guide label/body readable after autosize.
        worksheet.column_dimensions["A"].width = max(
            worksheet.column_dimensions["A"].width or 0, 14
        )
        worksheet.column_dimensions["B"].width = max(
            worksheet.column_dimensions["B"].width or 0, 28
        )

    def _find_refresh_command(self) -> Path | None:
        """Locate scripts/refresh_workbook.command by walking up from the workbook."""
        for parent in (self.output_path.resolve().parent, *self.output_path.resolve().parents):
            candidate = parent / "scripts" / "refresh_workbook.command"
            if candidate.is_file():
                return candidate
        return None

    def _write_refresh_control(self, overview: Worksheet, row_number: int) -> None:
        """Green clickable control that launches the Mac refresh script."""
        label = overview.cell(row=row_number, column=1, value="Update data")
        label.font = _REFRESH_LABEL_FONT

        action = overview.cell(row=row_number, column=2, value="Refresh from TWS")
        action.font = _REFRESH_FONT
        action.fill = _REFRESH_FILL
        action.alignment = Alignment(horizontal="center", vertical="center")

        command = self._find_refresh_command()
        if command is not None:
            # file:// URI so Excel for Mac can open the .command in Terminal.
            action.hyperlink = command.resolve().as_uri()
        else:
            action.value = "Run ./scripts/refresh_workbook.command"

        hint_row = row_number + 1
        hint = overview.cell(
            row=hint_row,
            column=1,
            value=(
                "Shortcut: closes this workbook, runs the IBKR snapshot, "
                "reopens the file. Requires TWS/Gateway API on."
            ),
        )
        hint.font = _GUIDE_BODY_FONT
        overview.merge_cells(
            start_row=hint_row, start_column=1, end_row=hint_row, end_column=2
        )

    def _write_overview(
        self,
        workbook: Workbook,
        data: dict[str, list[dict[str, Any]]],
    ) -> None:
        overview = workbook["IBKR_Overview"]
        overview.delete_rows(1, overview.max_row)

        data_start = self._write_sheet_guide(overview, "IBKR_Overview")

        generated_at = datetime.now().astimezone()
        errors = data.get("errors", [])
        error_kinds = {"info": 0, "warning": 0, "error": 0}
        for row in errors:
            error_kinds[row.get("kind", "warning")] = (
                error_kinds.get(row.get("kind", "warning"), 0) + 1
            )

        accounts = [
            str(row.get("account", "")).strip()
            for row in data.get("accounts", [])
            if row.get("account")
        ]
        account_label = ", ".join(accounts) if accounts else "(none)"

        overview_rows: list[tuple[str, Any]] = [
            ("IBKR Portfolio Workbook", ""),
            ("Generated", generated_at.isoformat(timespec="seconds")),
            ("Exporter mode", self.mode),
            ("Connection mode", "Read-only"),
            ("Accounts", account_label),
            ("Account summary rows", len(data.get("account_summary", []))),
            ("Cash / P&L per currency rows", len(data.get("account_values", []))),
            ("Positions", len(data.get("positions", []))),
            ("Portfolio rows", len(data.get("portfolio", []))),
            (
                "API messages",
                f"{len(errors)} "
                f"(info: {error_kinds['info']}, "
                f"warnings: {error_kinds['warning']}, "
                f"errors: {error_kinds['error']})",
            ),
        ]

        for offset, row in enumerate(overview_rows):
            row_number = data_start + offset
            left = overview.cell(row=row_number, column=1, value=row[0])
            right = overview.cell(row=row_number, column=2, value=row[1])
            left.font = _BODY_FONT
            right.font = _BODY_FONT

        title_row = data_start
        overview.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=2)
        overview.cell(row=title_row, column=1).font = _TITLE_FONT
        overview.cell(row=title_row, column=1).alignment = Alignment(
            horizontal="left", vertical="center"
        )

        # Blank spacer, then the in-sheet refresh control.
        control_row = data_start + len(overview_rows) + 1
        self._write_refresh_control(overview, control_row)

        overview.column_dimensions["A"].width = 32
        overview.column_dimensions["B"].width = 42
        overview.freeze_panes = f"A{data_start}"

    def _write_reconciliation(
        self,
        workbook: Workbook,
        data: dict[str, list[dict[str, Any]]],
    ) -> None:
        fiscal = extract_fiscal_quantities(self.output_path)
        rows = reconcile_quantities(
            live=data.get("positions", []),
            fiscal=fiscal,
            tolerance=self.qty_tolerance,
        )
        self._write_records(
            workbook,
            "IBKR_Reconciliacao",
            rows,
            "IBKRReconciliacaoTable",
            preferred_order=(
                "symbol",
                "qty_ibkr",
                "qty_fiscal",
                "delta",
                "status",
            ),
        )

    # -- public --

    def export(self, data: dict[str, list[dict[str, Any]]]) -> Path:
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = self._load_or_create_workbook()
        self._write_overview(workbook, data)

        self._write_records(
            workbook,
            "IBKR_Account_Summary",
            data.get("account_summary", []),
            "IBKRAccountSummaryTable",
            preferred_order=(
                "account",
                "tag",
                "value",
                "value_numeric",
                "currency",
                "request_id",
            ),
        )

        self._write_records(
            workbook,
            "IBKR_Cash_By_Currency",
            data.get("account_values", []),
            "IBKRAccountValuesTable",
            preferred_order=(
                "account",
                "currency",
                "metric",
                "value_numeric",
                "value",
            ),
        )

        self._write_records(
            workbook,
            "IBKR_Positions",
            data.get("positions", []),
            "IBKRPositionsTable",
            preferred_order=(
                "account",
                "symbol",
                "instrument_kind",
                "security_type",
                "currency",
                "exchange",
                "primary_exchange",
                "contract_id",
                "quantity",
                "average_cost",
            ),
        )

        self._write_records(
            workbook,
            "IBKR_Portfolio",
            data.get("portfolio", []),
            "IBKRPortfolioTable",
            preferred_order=(
                "account",
                "symbol",
                "instrument_kind",
                "security_type",
                "currency",
                "exchange",
                "contract_id",
                "quantity",
                "market_price",
                "market_value",
                "average_cost",
                "unrealized_pnl",
                "realized_pnl",
            ),
        )

        self._write_records(
            workbook,
            "IBKR_API_Messages",
            data.get("errors", []),
            "IBKRAPIMessagesTable",
            preferred_order=(
                "kind",
                "code",
                "message",
                "error_time_local",
                "request_id",
                "error_time",
                "advanced_rejection",
            ),
        )

        if self.mode == "tax_workbook":
            self._write_reconciliation(workbook, data)
            self._write_event_sheets(workbook)

        self._ensure_foreign_sheet_guides(workbook)

        workbook.save(self.output_path)
        return self.output_path

    def _ensure_foreign_sheet_guides(self, workbook: Workbook) -> None:
        """Add Description / How to read / Tip to preserved foreign tabs."""
        for sheet_name, guide in SHEET_GUIDES.items():
            if sheet_name in self.owned_sheets:
                continue
            if sheet_name not in workbook.sheetnames:
                continue
            ws = workbook[sheet_name]
            if ws["A1"].value == "Description":
                # Refresh text in place (unmerge first so writes stick).
                for r in range(1, 4):
                    for merged in list(ws.merged_cells.ranges):
                        if (
                            merged.min_row <= r <= merged.max_row
                            and merged.min_col <= 2 <= merged.max_col
                        ):
                            try:
                                ws.unmerge_cells(str(merged))
                            except ValueError:
                                pass
                ws["B1"] = guide["description"]
                ws["B2"] = guide["how_to_read"]
                ws["B3"] = guide["tip"]
                for r in range(1, 4):
                    ws.cell(row=r, column=1).font = _GUIDE_LABEL_FONT
                    body = ws.cell(row=r, column=2)
                    body.font = _GUIDE_BODY_FONT
                    body.alignment = Alignment(wrap_text=True, vertical="top")
                    ws.cell(row=r, column=1).fill = _GUIDE_FILL
                    body.fill = _GUIDE_FILL
                continue
            # Insert guide rows at the top without wiping existing content.
            ws.insert_rows(1, _GUIDE_ROW_COUNT)
            lines = (
                ("Description", guide["description"]),
                ("How to read", guide["how_to_read"]),
                ("Tip", guide["tip"]),
            )
            for index, (label, text) in enumerate(lines, start=1):
                label_cell = ws.cell(row=index, column=1, value=label)
                body_cell = ws.cell(row=index, column=2, value=text)
                label_cell.font = _GUIDE_LABEL_FONT
                body_cell.font = _GUIDE_BODY_FONT
                body_cell.alignment = Alignment(wrap_text=True, vertical="top")
                label_cell.fill = _GUIDE_FILL
                body_cell.fill = _GUIDE_FILL
                ws.row_dimensions[index].height = 72 if label == "Tip" else 40
            ws.row_dimensions[_GUIDE_ROW_COUNT].height = 8

    def _write_event_sheets(self, workbook: Workbook) -> None:
        from events_store import load_events
        from promote_events import derive_positions_from_events

        events: list[dict[str, Any]] = []
        if self.events_file and self.events_file.exists():
            events = load_events(self.events_file)

        staging_rows = [
            {
                "event_id": e.get("event_id"),
                "date": e.get("date"),
                "symbol": e.get("symbol"),
                "tipo_evento": e.get("tipo_evento"),
                "quantity": e.get("quantity"),
                "price_usd": e.get("price_usd"),
                "ir_retido_usd": e.get("ir_retido_usd"),
                "ptax": e.get("ptax"),
                "promoted": bool(e.get("promoted")),
                "source_file": e.get("source_file"),
                "observacoes": e.get("observacoes"),
            }
            for e in sorted(
                events,
                key=lambda x: (str(x.get("date", "")), str(x.get("symbol", ""))),
            )
        ]
        self._write_records(
            workbook,
            "IBKR_Eventos_Staging",
            staging_rows,
            "IBKREventosStagingTable",
            preferred_order=(
                "event_id",
                "date",
                "symbol",
                "tipo_evento",
                "quantity",
                "price_usd",
                "ir_retido_usd",
                "ptax",
                "promoted",
                "source_file",
                "observacoes",
            ),
        )
        self._write_records(
            workbook,
            "IBKR_Posicao_From_Events",
            derive_positions_from_events(events),
            "IBKRPosicaoFromEventsTable",
            preferred_order=(
                "symbol",
                "quantity",
                "avg_cost_usd",
                "avg_cost_brl",
                "status",
            ),
        )


__all__ = [
    "ExcelExporter",
    "OWNED_SHEETS_STANDALONE",
    "OWNED_SHEETS_TAX",
    "LEGACY_OWNED_SHEETS",
    "extract_fiscal_quantities",
]
