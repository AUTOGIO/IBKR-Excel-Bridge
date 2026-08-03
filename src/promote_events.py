"""Promote canonical events into Registro_Real (append-only)."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from events_store import load_events, save_events


_FORMULA_COLS = ("H", "I", "J", "K", "M", "N", "O")
_EVENT_ID_RE = re.compile(r"event_id=([0-9a-f]+)")


def _assert_unlocked(path: Path) -> None:
    lock = path.parent / f"~${path.name}"
    if lock.exists():
        raise PermissionError(
            f"Workbook appears open in Excel ({lock.name}). Close Excel and retry."
        )


def _last_data_row(ws) -> int:
    last = 4
    for row_idx in range(5, ws.max_row + 1):
        if ws.cell(row_idx, 1).value not in (None, ""):
            last = row_idx
    return last


def _existing_event_ids(ws) -> set[str]:
    found: set[str] = set()
    for row_idx in range(5, ws.max_row + 1):
        obs = ws.cell(row_idx, 12).value  # L
        if not obs:
            continue
        match = _EVENT_ID_RE.search(str(obs))
        if match:
            found.add(match.group(1))
    return found


def _retarget_formula(formula: str, from_row: int, to_row: int) -> str:
    """Replace row indices that refer to from_row with to_row (sheet-local cells)."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    pattern = re.compile(r"(\$?)([A-Z]+)(\$?)(\d+)")

    def sub(m: re.Match[str]) -> str:
        col_abs, col, row_abs, row_s = m.group(1), m.group(2), m.group(3), m.group(4)
        row = int(row_s)
        if row == from_row:
            row = to_row
        return f"{col_abs}{col}{row_abs}{row}"

    return pattern.sub(sub, formula)


def _copy_formulas(ws, template_row: int, target_row: int) -> None:
    for col in _FORMULA_COLS:
        value = ws[f"{col}{template_row}"].value
        if isinstance(value, str) and value.startswith("="):
            ws[f"{col}{target_row}"] = _retarget_formula(value, template_row, target_row)
        elif value is not None:
            # ArrayFormula or other — try string
            ws[f"{col}{target_row}"] = value


def promote_events_to_workbook(
    workbook_path: Path,
    events_path: Path,
) -> dict[str, Any]:
    """Append unpromoted events to Registro_Real. Returns summary stats."""
    if not workbook_path.exists():
        raise FileNotFoundError(f"Tax workbook not found: {workbook_path}")
    _assert_unlocked(workbook_path)

    events = load_events(events_path)
    pending = [e for e in events if not e.get("promoted")]
    if not pending:
        return {"appended": 0, "skipped": 0, "pending": 0}

    wb = load_workbook(workbook_path)
    if "Registro_Real" not in wb.sheetnames:
        raise ValueError("Workbook missing Registro_Real sheet")
    ws = wb["Registro_Real"]

    existing_ids = _existing_event_ids(ws)
    last_row = _last_data_row(ws)
    template_row = last_row if last_row >= 5 else 5
    appended = 0
    skipped = 0

    pending_sorted = sorted(
        pending, key=lambda e: (e.get("date", ""), e.get("symbol", ""), e.get("event_id", ""))
    )

    for event in pending_sorted:
        eid = str(event["event_id"])
        if eid in existing_ids:
            event["promoted"] = True
            skipped += 1
            continue
        target = last_row + 1
        last_row = target
        ws.cell(target, 1, value=_as_date(event["date"]))
        ws.cell(target, 2, value=event["symbol"])
        ws.cell(target, 3, value=event["tipo_evento"])
        ws.cell(target, 4, value=float(event["quantity"]))
        ws.cell(target, 5, value=event.get("price_usd"))
        ws.cell(target, 6, value=event.get("ptax"))
        ws.cell(target, 7, value=event.get("ir_retido_usd"))
        ws.cell(target, 12, value=event.get("observacoes"))
        _copy_formulas(ws, template_row, target)
        existing_ids.add(eid)
        event["promoted"] = True
        appended += 1
        template_row = target  # next formulas expand from newest row

    # Persist promoted flags
    by_id = {str(e["event_id"]): e for e in events}
    for event in pending_sorted:
        by_id[str(event["event_id"])] = event
    save_events(events_path, list(by_id.values()))
    wb.save(workbook_path)
    return {
        "appended": appended,
        "skipped": skipped,
        "pending": len(pending),
        "last_row": last_row,
    }


def _as_date(value: str):
    from datetime import date, datetime

    if isinstance(value, (date, datetime)):
        return value
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


def write_staging_sheets(
    workbook_path: Path,
    events: list[dict[str, Any]],
) -> None:
    """Rewrite IBKR_Eventos_Staging and IBKR_Posicao_From_Events on the tax workbook."""
    from excel_exporter import ExcelExporter

    if not workbook_path.exists():
        raise FileNotFoundError(f"Tax workbook not found: {workbook_path}")
    _assert_unlocked(workbook_path)

    exporter = ExcelExporter(workbook_path, mode="tax_workbook")
    # Reuse private writers via a thin public helper on the exporter instance
    wb = load_workbook(workbook_path)
    for name in ("IBKR_Eventos_Staging", "IBKR_Posicao_From_Events"):
        if name in wb.sheetnames:
            del wb[name]

    staging_rows = [
        {
            "event_id": e.get("event_id"),
            "date": e.get("date"),
            "symbol": e.get("symbol"),
            "tipo_evento": e.get("tipo_evento"),
            "quantity": e.get("quantity"),
            "price_usd": e.get("price_usd"),
            "ir_retido_usd": e.get("ir_retido_usd"),
            "ptax": e.get("ptax"),
            "promoted": bool(e.get("promoted")),
            "source_file": e.get("source_file"),
            "observacoes": e.get("observacoes"),
        }
        for e in sorted(
            events,
            key=lambda x: (str(x.get("date", "")), str(x.get("symbol", ""))),
        )
    ]
    exporter._write_records(  # noqa: SLF001 — shared table writer
        wb,
        "IBKR_Eventos_Staging",
        staging_rows,
        "IBKREventosStagingTable",
        preferred_order=(
            "event_id",
            "date",
            "symbol",
            "tipo_evento",
            "quantity",
            "price_usd",
            "ir_retido_usd",
            "ptax",
            "promoted",
            "source_file",
            "observacoes",
        ),
    )

    positions = derive_positions_from_events(events)
    exporter._write_records(  # noqa: SLF001
        wb,
        "IBKR_Posicao_From_Events",
        positions,
        "IBKRPosicaoFromEventsTable",
        preferred_order=(
            "symbol",
            "quantity",
            "avg_cost_usd",
            "avg_cost_brl",
            "status",
        ),
    )
    wb.save(workbook_path)


def derive_positions_from_events(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Simple moving average inventory in USD (BRL when ptax present)."""
    state: dict[str, dict[str, float]] = {}
    for event in sorted(events, key=lambda e: (str(e.get("date", "")), str(e.get("event_id", "")))):
        symbol = str(event.get("symbol", "")).upper()
        if not symbol:
            continue
        tipo = event.get("tipo_evento")
        qty = float(event.get("quantity") or 0)
        price = event.get("price_usd")
        ptax = event.get("ptax")
        slot = state.setdefault(
            symbol, {"qty": 0.0, "cost_usd": 0.0, "cost_brl": 0.0}
        )
        if tipo == "Compra" and price is not None:
            slot["cost_usd"] += qty * float(price)
            if ptax is not None:
                slot["cost_brl"] += qty * float(price) * float(ptax)
            slot["qty"] += qty
        elif tipo == "Venda":
            if slot["qty"] <= 0:
                continue
            avg_usd = slot["cost_usd"] / slot["qty"] if slot["qty"] else 0.0
            avg_brl = slot["cost_brl"] / slot["qty"] if slot["qty"] else 0.0
            sell = min(qty, slot["qty"])
            slot["qty"] -= sell
            slot["cost_usd"] -= sell * avg_usd
            slot["cost_brl"] -= sell * avg_brl
        # Rendimento: no position change

    rows: list[dict[str, Any]] = []
    for symbol, slot in sorted(state.items()):
        qty = slot["qty"]
        if abs(qty) < 1e-12:
            continue
        avg_usd = slot["cost_usd"] / qty if qty else None
        avg_brl = slot["cost_brl"] / qty if qty and slot["cost_brl"] else None
        rows.append(
            {
                "symbol": symbol,
                "quantity": qty,
                "avg_cost_usd": avg_usd,
                "avg_cost_brl": avg_brl,
                "status": "DERIVED_FROM_EVENTS",
            }
        )
    return rows


__all__ = [
    "promote_events_to_workbook",
    "write_staging_sheets",
    "derive_positions_from_events",
]
